import streamlit as st
import pandas as pd
import os
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import io
import csv

def identify_container_column(df):
    """
    Identify the container number column in the dataframe based on known terminology.
    """
    if df is None:
        return None

    possible_column_names = [
        'CONTAINER NUMBER', 'CONTAINER', 'CONTAINER NO', 'CONTAINER_NUMBER',
        'container', 'container_number', 'container_no', 'containerno', 'container_id',
        'Unit No', 'UNIT NO', 'unit no', 'unit_no', 'unitno', 'UNIT_NO'
    ]

    # Exact matches first
    for col_name in possible_column_names:
        if col_name in df.columns:
            return col_name

    # Then partial matches
    for col in df.columns:
        if 'container' in col.lower() or 'unit' in col.lower():
            return col

    return None

def identify_columns(df, system_type):
    """
    Identify relevant columns in the dataframes based on the system type.
    
    Args:
        df: The dataframe to analyze
        system_type: Either 'TOPS' or 'Cyman'
        
    Returns:
        dict: A dictionary of identified column names
    """
    identified_columns = {}
    
    if df is None:
        return identified_columns
    
    if system_type == 'TOPS':
        # Identify container column
        container_col = identify_container_column(df)
        identified_columns['container_col'] = container_col
        
        # Identify status column
        for col in df.columns:
            if any(status in str(col).lower() for status in ['status', 'state', 'progress']):
                identified_columns['status_col'] = col
                break
        
        # Identify location column
        for col in df.columns:
            if any(loc in str(col).lower() for loc in ['location', 'unload', 'terminal']):
                identified_columns['location_col'] = col
                break
    
    elif system_type == 'Cyman':
        # Identify container column
        container_col = identify_container_column(df)
        identified_columns['container_col'] = container_col
        
        # Identify activity column
        for col in df.columns:
            if any(act in str(col).lower() for act in ['activity', 'status', 'standard']):
                identified_columns['activity_col'] = col
                break
    
    return identified_columns

def format_excel_workbook(wb, row_count):
    """
    Apply formatting to the Excel workbook (in memory).
    """
    ws = wb.active

    # Define styles
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    missing_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    present_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    center_alignment = Alignment(horizontal='center', vertical='center')

    # Format header row
    for col in range(1, 4):  # Columns A, B, C
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment

    # Format data rows
    for row in range(2, row_count + 2):  # Skip header row
        # Container Number column
        ws.cell(row=row, column=1).alignment = center_alignment

        # Format CYMAN column (column B)
        cyman_cell = ws.cell(row=row, column=2)
        cyman_cell.alignment = center_alignment
        if cyman_cell.value == 'Missing':
            cyman_cell.fill = missing_fill
            cyman_cell.value = '❌'
        else:
            cyman_cell.fill = present_fill
            cyman_cell.value = '✓'

        # Format TOPS column (column C)
        tops_cell = ws.cell(row=row, column=3)
        tops_cell.alignment = center_alignment
        if tops_cell.value == 'Missing':
            tops_cell.fill = missing_fill
            tops_cell.value = '❌'
        else:
            tops_cell.fill = present_fill
            tops_cell.value = '✓'

    # Set column widths
    ws.column_dimensions[get_column_letter(1)].width = 20
    ws.column_dimensions[get_column_letter(2)].width = 12
    ws.column_dimensions[get_column_letter(3)].width = 12

    # Add a title row
    ws.insert_rows(1)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = "Container Number Comparison Report"
    title_cell.font = Font(size=14, bold=True)
    ws.merge_cells('A1:C1')
    title_cell.alignment = Alignment(horizontal='center')

    # Add terminology note
    ws.insert_rows(2)
    terminology_cell = ws.cell(row=2, column=1)
    terminology_cell.value = "Note: TOPS 'Container Number' = Cyman 'Unit No'"
    terminology_cell.font = Font(italic=True)
    ws.merge_cells('A2:C2')
    terminology_cell.alignment = Alignment(horizontal='center')

    # Add timestamp
    ws.insert_rows(3)
    date_cell = ws.cell(row=3, column=1)
    date_cell.value = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws.merge_cells('A3:C3')
    date_cell.alignment = Alignment(horizontal='center')

    # Add summary at the bottom
    ws.append([])  # Empty row
    summary_row = ws.max_row + 1
    summary_cell = ws.cell(row=summary_row, column=1)
    summary_cell.value = f"Total mismatches found: {row_count}"
    summary_cell.font = Font(bold=True)
    ws.merge_cells(f'A{summary_row}:C{summary_row}')

def format_for_display(results_df):
    """
    Format the results DataFrame for on-screen display with the same style as the Excel output.
    """
    # Create a copy of the dataframe for display
    display_df = results_df.copy()
    
    # Replace values with symbols for display
    display_df = display_df.replace({'Missing': '❌', 'Present': '✓'})
    
    return display_df

def compare_container_spreadsheets(tops_file, cyman_file, tops_container_col=None, cyman_container_col=None, check_single_boxes=False):
    """
    Compare container numbers between TOPS and Cyman spreadsheets with enhanced logic.
    Returns a tuple containing:
    - BytesIO for Excel output
    - DataFrame for display
    - StringIO for CSV output
    """
    # Load spreadsheets
    try:
        tops_df = pd.read_excel(tops_file)
        cyman_df = pd.read_excel(cyman_file)
    except Exception as e:
        st.error(f"Error loading spreadsheets: {e}")
        return None, None, None

    # Identify columns in both spreadsheets
    tops_columns = identify_columns(tops_df, 'TOPS')
    cyman_columns = identify_columns(cyman_df, 'Cyman')
    
    # Auto-detect container columns if not provided
    if tops_container_col is None:
        tops_container_col = tops_columns.get('container_col')
        if tops_container_col:
            st.write(f"TOPS container column detected: {tops_container_col}")
        else:
            st.error("Could not automatically detect container number column in TOPS. Available columns:")
            st.write(list(tops_df.columns))
            return None, None, None

    if cyman_container_col is None:
        cyman_container_col = cyman_columns.get('container_col')
        if cyman_container_col:
            st.write(f"Cyman container column detected: {cyman_container_col}")
        else:
            st.error("Could not automatically detect container number column in Cyman. Available columns:")
            st.write(list(cyman_df.columns))
            return None, None, None
            
    # Initialize columns for status and location/activity
    tops_status_col = tops_columns.get('status_col', None)
    tops_location_col = tops_columns.get('location_col', None)
    cyman_activity_col = cyman_columns.get('activity_col', None)
    
    # If we couldn't detect columns automatically, look for them by position based on the sample data
    if tops_status_col is None and len(tops_df.columns) >= 1:
        tops_status_col = tops_df.columns[0]  # First column in TOPS sample
        st.write(f"Using first column as TOPS status: {tops_status_col}")
    
    if tops_location_col is None and len(tops_df.columns) >= 3:
        tops_location_col = tops_df.columns[2]  # Third column in TOPS sample
        st.write(f"Using third column as TOPS location: {tops_location_col}")
    
    if cyman_activity_col is None and len(cyman_df.columns) >= 6:
        cyman_activity_col = cyman_df.columns[5]  # Sixth column in Cyman sample
        st.write(f"Using sixth column as Cyman activity: {cyman_activity_col}")

    try:
        # Convert container columns to strings and strip whitespace
        tops_df[tops_container_col] = tops_df[tops_container_col].astype(str).str.strip()
        cyman_df[cyman_container_col] = cyman_df[cyman_container_col].astype(str).str.strip()
        
        # Filter TOPS data to only include "Complete" or "In Progress" in Status Name column
        # And only include "James Kemball Holding Centre" in Unload Location column
        if tops_status_col:
            tops_df = tops_df[tops_df[tops_status_col].astype(str).str.lower().isin(['complete', 'in progress'])]
        
        if tops_location_col:
            tops_df = tops_df[tops_df[tops_location_col].astype(str).str.lower() == 'james kemball holding centre']
        
        # Create dictionaries for faster lookups
        tops_containers = set(tops_df[tops_container_col])
        cyman_containers = set(cyman_df[cyman_container_col])
        
        tops_data = {row[tops_container_col]: row for _, row in tops_df.iterrows()}
        cyman_data = {row[cyman_container_col]: row for _, row in cyman_df.iterrows()}
    except KeyError as e:
        st.error(f"Error: Column not found - {e}")
        return None, None, None

    # Process containers based on the enhanced requirements
    results = []
    
    # Process containers in TOPS but not in Cyman
    for container in tops_containers - cyman_containers:
        results.append({
            'CONTAINER NUMBER': container,
            'CYMAN': 'Missing',
            'TOPS': 'Present'
        })
    
    # Process containers in Cyman but not in TOPS
    for container in cyman_containers - tops_containers:
        results.append({
            'CONTAINER NUMBER': container,
            'CYMAN': 'Present',
            'TOPS': 'Missing'
        })
    
    # Check for single boxes across both yards if option is selected
    if check_single_boxes:
        # Add logic to find containers that occur only once across both systems
        # This is a simple implementation looking for unique container identifiers
        all_containers = list(tops_containers) + list(cyman_containers)
        from collections import Counter
        container_counts = Counter(all_containers)
        single_boxes = [container for container, count in container_counts.items() if count == 1]
        
        for container in single_boxes:
            # Check if it's not already in our results (which would mean it's missing from one system)
            if container not in tops_containers and container not in cyman_containers:
                if container in tops_containers:
                    results.append({
                        'CONTAINER NUMBER': container,
                        'CYMAN': 'Missing',
                        'TOPS': 'Present'
                    })
                else:
                    results.append({
                        'CONTAINER NUMBER': container,
                        'CYMAN': 'Present',
                        'TOPS': 'Missing'
                    })
    
    # Sort the results alphabetically by container number
    results.sort(key=lambda x: x['CONTAINER NUMBER'])
    
    # Create DataFrame from results
    results_df = pd.DataFrame(results)
    
    if results_df.empty:
        st.info("No mismatches found based on the specified criteria.")
        empty_df = pd.DataFrame(columns=['CONTAINER NUMBER', 'CYMAN', 'TOPS'])
        output_buffer = io.BytesIO()
        empty_df.to_excel(output_buffer, index=False)
        output_buffer.seek(0)
        
        csv_buffer = io.StringIO()
        empty_df.to_csv(csv_buffer, index=False)
        csv_buffer.seek(0)
        
        return output_buffer, empty_df, csv_buffer

    # Write results to an in-memory Excel file
    output_buffer = io.BytesIO()
    results_df.to_excel(output_buffer, index=False)
    output_buffer.seek(0)

    # Load workbook from the BytesIO buffer
    try:
        wb = openpyxl.load_workbook(output_buffer)
    except Exception as e:
        st.error(f"Error processing workbook: {e}")
        return None, None, None

    # Apply formatting
    format_excel_workbook(wb, len(results_df))

    final_buffer = io.BytesIO()
    wb.save(final_buffer)
    final_buffer.seek(0)

    # Create CSV buffer
    csv_buffer = io.StringIO()
    results_df.to_csv(csv_buffer, index=False)
    csv_buffer.seek(0)

    st.success("Comparison complete!")
    return final_buffer, results_df, csv_buffer

def main():
    st.title("Container Spreadsheet Comparison Tool")

    # Upload file widgets
    tops_file = st.file_uploader("Upload TOPS spreadsheet", type=["xlsx", "xls"])
    cyman_file = st.file_uploader("Upload Cyman spreadsheet", type=["xlsx", "xls"])

    # Text inputs for container column names (optional)
    tops_col = st.text_input("TOPS container column name (leave blank for auto-detection)")
    if tops_col.strip() == "":
        tops_col = None

    cyman_col = st.text_input("Cyman container column name (leave blank for auto-detection)")
    if cyman_col.strip() == "":
        cyman_col = None
        
    # Option to check for single boxes
    check_single_boxes = st.checkbox("Check for single boxes across both yards")
    
    # Output format selection
    output_format = st.radio("Select output format:", ["View on screen", "Download as Excel", "Download as CSV"])

    if st.button("Run Comparison"):
        if not tops_file or not cyman_file:
            st.error("Please upload both TOPS and Cyman spreadsheets.")
        else:
            # Fixed the return value unpacking
            excel_buffer, result_df, csv_buffer = compare_container_spreadsheets(
                tops_file, cyman_file, tops_col, cyman_col, check_single_boxes
            )
            
            if excel_buffer is not None and result_df is not None and csv_buffer is not None:
                if output_format == "View on screen":
                    # Display results on screen with the same formatting as the Excel/CSV
                    st.subheader("Container Number Comparison Report")
                    st.caption("Note: TOPS 'Container Number' = Cyman 'Unit No'")
                    st.caption(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
                    
                    # Format the dataframe for display
                    display_df = format_for_display(result_df)
                    
                    # Use CSS to style the dataframe similar to the Excel output
                    st.markdown("""
                    <style>
                    .dataframe th {
                        background-color: #1F4E78;
                        color: white;
                        text-align: center;
                    }
                    .dataframe td {
                        text-align: center;
                    }
                    </style>
                    """, unsafe_allow_html=True)
                    
                    st.dataframe(display_df)
                    st.write(f"Total mismatches found: {len(result_df)}")
                    
                elif output_format == "Download as Excel":
                    st.download_button(
                        label="Download Comparison Report (Excel)",
                        data=excel_buffer,
                        file_name="container_comparison_report.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    
                elif output_format == "Download as CSV":
                    st.download_button(
                        label="Download Comparison Report (CSV)",
                        data=csv_buffer.getvalue(),
                        file_name="container_comparison_report.csv",
                        mime="text/csv"
                    )

if __name__ == "__main__":
    main()
